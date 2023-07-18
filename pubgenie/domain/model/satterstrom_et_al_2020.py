from openpyxl import load_workbook
from html import escape

def gene_query(user_input, filename, worksheet_index):
    # Now handle the Excel file
    workbook = load_workbook(filename=filename, read_only=True)
    # Focus on the specific worksheet
    sheet = workbook.worksheets[worksheet_index]

    # Get the data
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    headers = data[0]
    data_rows = [row for row in data[1:] if user_input in row]

    table_html = '<h2><a href="https://doi.org/10.1016/j.cell.2019.12.036" target="_blank">Satterstrom <i>et al.</i> Cell 2020</a></h2>'
    table_html += '<table>\n'
    table_html += '<tr>' + ''.join(f'<th>{escape(str(header))}</th>' for header in headers) + '</tr>\n'
    for row in data_rows:
        table_html += '<tr>' + ''.join(f'<td>{escape(str(cell))}</td>' for cell in row) + '</tr>\n'
    table_html += '</table>'

    return table_html
