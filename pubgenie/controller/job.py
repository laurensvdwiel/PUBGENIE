import subprocess
from openpyxl import load_workbook
from html import escape

def execute_query(user_input):
    command = f"head -1 /data/41586_2020_2832_MOESM3_ESM.txt && zgrep {user_input} /data/41586_2020_2832_MOESM3_ESM.txt"
    process = subprocess.Popen(command, stdout=subprocess.PIPE, shell=True)
    output, error = process.communicate()

    if error:
        print(f"Error: {error}")
        return f"Error: {escape(error.decode('utf-8'))}"

    lines = output.decode('utf-8').split('\n')
    headers = lines[0].split('\t')
    data_rows = [line.split('\t') for line in lines[1:] if line]

    table_html = '<h2><a href="https://www.nature.com/articles/s41586-020-2832-5" target="_blank">Kaplanis <i>et al.</i> Nature 2020</a></h2>'
    table_html += '<table>\n'
    table_html += '<tr>' + ''.join(f'<th>{escape(header)}</th>' for header in headers) + '</tr>\n'
    for row in data_rows:
        table_html += '<tr>' + ''.join(f'<td>{escape(cell)}</td>' for cell in row) + '</tr>\n'
    table_html += '</table>'

    # Now handle the Excel file
    workbook = load_workbook(filename='/data/1-s2.0-S0092867419313984-mmc1.xlsx', read_only=True)
    sheet = workbook.worksheets[4]  # Get the 5th sheet

    # Get the data
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    headers = data[0]
    data_rows = [row for row in data[1:] if user_input in row]

    table_html += '<h2><a href="https://doi.org/10.1016/j.cell.2019.12.036" target="_blank">Satterstrom <i>et al.</i> Cell 2020</a></h2>'
    table_html += '<table>\n'
    table_html += '<tr>' + ''.join(f'<th>{escape(str(header))}</th>' for header in headers) + '</tr>\n'
    for row in data_rows:
        table_html += '<tr>' + ''.join(f'<td>{escape(str(cell))}</td>' for cell in row) + '</tr>\n'
    table_html += '</table>'

    return table_html
